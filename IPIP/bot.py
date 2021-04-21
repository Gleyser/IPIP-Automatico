import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class bot():
    def __init__(self):
        self.drive = webdriver.Chrome(executable_path=r"C:\Users\gleys\Documents\GitHub\IPIP-Automatico\IPIP\chromedriver.exe")
        self.wb = load_workbook(filename='dados.xlsx')
        self.ws = self.wb['Dados']

    def preencherForms(self):
        drive = self.drive

        for pessoa in range (2, 121):
            drive.get('https://www.personal.psu.edu/~j5j/IPIP/ipipneo120.htm')
            time.sleep(10)
            yes1 = drive.find_element_by_xpath('/html/body/form/table[1]/tbody/tr/td/input')
            yes1.click()

            yes2 = drive.find_element_by_xpath('/html/body/form/table[2]/tbody/tr/td/input')
            yes2.click()

            botaoSend = drive.find_element_by_xpath('/html/body/form/b[2]/p/input')
            botaoSend.click()
            time.sleep(10)
            # esperar clicar no botao de aceitar sem seguranca
            
            botaoSendSeguranca = drive.find_element_by_xpath('//*[@id="proceed-button"]')
            botaoSendSeguranca.click()
            time.sleep(20)

            # linha da planilha 
            linhaDaPessoa = pessoa

            # nome
            nome = drive.find_element_by_xpath('/html/body/form/p[2]/input')
            nome.send_keys(self.ws[linhaDaPessoa][0].value)

            # sexo
            sexo = self.ws[linhaDaPessoa][2].value
            if (sexo == "Masculino"):
                sexoMasculino = drive.find_element_by_xpath('/html/body/form/p[3]/table/tbody/tr/td[2]/input')
                sexoMasculino.click()
            
            elif (sexo == "Feminino"):
                sexoFeminino = drive.find_element_by_xpath('/html/body/form/p[3]/table/tbody/tr/td[3]/input')
                sexoFeminino.click()

            # idade
            idade = drive.find_element_by_xpath('/html/body/form/p[4]/input')
            idade.send_keys(self.ws[linhaDaPessoa][3].value)

            # pais
            pais = drive.find_element_by_xpath('/html/body/form/p[6]/select')
            pais.send_keys('Brazil')

            # Parte 1 do formulario
            for numeroQuestaoFormulario in range(1, 61):
                numeroQuestaoExcel = numeroQuestaoFormulario + 3
                respostaDaPessoa = self.ws[linhaDaPessoa][numeroQuestaoExcel].value
                qfinal = ""

                if (respostaDaPessoa == "Muito Impreciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/p[6]/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[3]/center/input')
                    
                elif (respostaDaPessoa == "Moderadamente Impreciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/p[6]/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[4]/center/input')

                elif (respostaDaPessoa == "Nem preciso Nem impreciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/p[6]/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[5]/center/input')

                elif (respostaDaPessoa == "Moderadamente Preciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/p[6]/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[6]/center/input')

                elif (respostaDaPessoa == "Muito Preciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/p[6]/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[7]/center/input')      
                
                qfinal.click()

            botaoEnviarParte1 = drive.find_element_by_xpath('/html/body/form/p[7]/input')
            botaoEnviarParte1.click()        
            
            # esperar a transicao entre as telas
            time.sleep(20)
            
            # Parte 2 do formulario
            for numeroQuestaoFormulario in range(1, 61):
                numeroQuestaoExcel = numeroQuestaoFormulario + 63
                respostaDaPessoa = self.ws[linhaDaPessoa][numeroQuestaoExcel].value
                qfinal = ""

                if (respostaDaPessoa == "Muito Impreciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[3]/center/input')          
                    
                elif (respostaDaPessoa == "Moderadamente Impreciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[4]/center/input')  

                elif (respostaDaPessoa == "Nem preciso Nem impreciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[5]/center/input')  

                elif (respostaDaPessoa == "Moderadamente Preciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[6]/center/input')  

                elif (respostaDaPessoa == "Muito Preciso"):
                    qfinal = drive.find_element_by_xpath('/html/body/form/table/tbody/tr['+str(numeroQuestaoFormulario)+']/td[7]/center/input')                
                
                qfinal.click()  

            # buffer
            time.sleep(5)  

            botaoEnviarParte2 = drive.find_element_by_xpath('/html/body/form/p[3]/input')
            botaoEnviarParte2.click()             

            # buffer
            time.sleep(10)            
            
            WebDriverWait(drive, 500).until(EC.title_is("IPIP-NEO Narrative Report"))
            texto = drive.find_element_by_xpath('/html/body').text
            time.sleep(10)
            arquivo = open(str(linhaDaPessoa) + "-" + str(self.ws[linhaDaPessoa][0].value) + ".txt", "a")
            arquivo.write(str(texto))   
            arquivo.close()    

            # esperar a transicao entre as telas
            time.sleep(3)    
        

bot = bot()
bot.preencherForms()

