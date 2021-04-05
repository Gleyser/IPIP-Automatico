from openpyxl import load_workbook

wb = load_workbook(filename='dados.xlsx')
ws = wb['Dados']

for row in ws.rows:
    for cell in row:
        print(cell.value)
        break
    break

print(ws[1][0].value)

for i in range(1, 61):
    print(i)
